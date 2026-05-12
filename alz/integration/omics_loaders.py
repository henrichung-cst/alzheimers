"""Bulk omics loaders for the Incytr factorial fixture.

Parses the Song lab merged-labeled xlsx files for proteomics (pr) and
phosphoproteomics (ps, py), produces gene × animal matrices keyed by
the transcript-side animal_id so they line up with the design matrix.

xlsx layout (all three files): descriptive sample headers in row 1
(e.g. ``1_C198(L)_M_2mo_WT_P1_128N``), machine column names in row 2
(e.g. ``plex1_128n_sn_mean``). Sample columns start after a fixed
preamble of metadata columns (different across pr/ps/py).

Animal-ID canonicalization: transcript fixtures use a zero-padded
form (``D092``), proteomics descriptive labels use unpadded
(``D92``). Canonical = strip leading zeros after the first letter
prefix.

Phospho site → gene aggregator is ``sum``: symmetric to bulk
proteomics totals, tolerates missing sites, interpretable as total
phospho-occupancy fold change.
"""

from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass
from typing import Callable

import numpy as np
import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config as _config  # noqa: E402

CANON_ID_RE = re.compile(r"^([A-Za-z]+)0*(\d+)")
PROTEIN_ID_KEEP_PREFIX = "ENSMUSP"


@dataclass(frozen=True)
class OmicsSchema:
    """Per-layer xlsx parsing schema."""
    layer: str
    path: str
    gene_col_idx: int            # 0-based column index of the gene symbol
    first_sample_col: int        # 0-based column index of the first sample column
    protein_id_col_idx: int      # 0-based column index of the protein_id (for ENSMUSP filter)
    aggregator: str              # "mean" for protein-level, "sum" for site-level


OMICS_SCHEMAS: dict[str, OmicsSchema] = {
    "pr": OmicsSchema(
        layer="pr",
        path=_config.SONG_PROTEIN_QUANT_FILE,
        gene_col_idx=1,            # "Gene Symbol"
        first_sample_col=26,       # plex1_126_sn_mean
        protein_id_col_idx=0,      # "protein_id"
        aggregator="mean",         # mostly unique-per-gene; mean handles isoform dupes
    ),
    "ps": OmicsSchema(
        layer="ps",
        path=_config.SONG_IMAC_SITEQUANT_FILE,
        gene_col_idx=2,
        first_sample_col=41,
        protein_id_col_idx=1,
        aggregator="sum",          # site → gene collapse
    ),
    "py": OmicsSchema(
        layer="py",
        path=_config.SONG_PY_SITEQUANT_FILE,
        gene_col_idx=1,
        first_sample_col=106,
        protein_id_col_idx=0,
        aggregator="sum",          # site → gene collapse
    ),
}


def canonical_animal_id(label: str | None) -> str | None:
    """Extract canonical animal ID (letter prefix + unpadded digits).

    ``D092`` → ``D92``; ``C198`` → ``C198``; ``E50`` → ``E50``.
    Returns None if no animal-like pattern is found.
    """
    if not isinstance(label, str):
        return None
    m = CANON_ID_RE.search(label)
    if not m:
        return None
    return f"{m.group(1).upper()}{int(m.group(2))}"


def extract_animal_from_descriptive(label: str | None) -> str | None:
    """Parse a Song proteomics descriptive header.

    Format: ``<idx>_<ANIMAL>(<L|R|N>)_M_<age>_<geno>_P<plex>_<channel>``.
    Returns canonical animal ID or None if it's a Ref_Pool / unparseable.
    """
    if not isinstance(label, str):
        return None
    if label.startswith("Ref_Pool"):
        return None
    parts = label.split("_", 2)
    if len(parts) < 2:
        return None
    return canonical_animal_id(parts[1])


def _load_xlsx_rows(path: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _build_animal_column_map(
    descriptive_row: tuple, schema: OmicsSchema
) -> dict[int, str]:
    out: dict[int, str] = {}
    for i in range(schema.first_sample_col, len(descriptive_row)):
        animal = extract_animal_from_descriptive(descriptive_row[i])
        if animal is not None:
            out[i] = animal
    return out


def _is_quantifiable_row(row: tuple, schema: OmicsSchema) -> bool:
    protein_id = row[schema.protein_id_col_idx]
    if not isinstance(protein_id, str) or not protein_id.startswith(PROTEIN_ID_KEEP_PREFIX):
        return False
    gene = row[schema.gene_col_idx]
    return isinstance(gene, str) and gene.strip() != ""


_AGGREGATORS: dict[str, Callable[[pd.core.groupby.DataFrameGroupBy], pd.DataFrame]] = {
    "mean": lambda g: g.mean(numeric_only=True),
    # min_count=1 preserves all-NaN semantics: a gene with NaN at every site
    # stays NaN per animal instead of collapsing to 0.
    "sum": lambda g: g.sum(min_count=1),
}


def load_gene_matrix(schema: OmicsSchema) -> pd.DataFrame:
    """Load one layer's xlsx and collapse to a genes × animals matrix
    using the schema's aggregator."""
    rows = _load_xlsx_rows(schema.path)
    descriptive_row = rows[0]
    col_to_animal = _build_animal_column_map(descriptive_row, schema)
    if not col_to_animal:
        raise RuntimeError(f"{schema.layer}: no animal columns parsed from descriptive row")

    sample_cols = sorted(col_to_animal.keys())
    animals_for_cols = [col_to_animal[c] for c in sample_cols]

    records: list[dict] = []
    for r in rows[2:]:  # skip descriptive (row 1) + machine header (row 2)
        if not _is_quantifiable_row(r, schema):
            continue
        rec: dict = {"gene": r[schema.gene_col_idx]}
        for animal, c in zip(animals_for_cols, sample_cols):
            v = r[c]
            rec[animal] = v if isinstance(v, (int, float)) else np.nan
        records.append(rec)

    df = pd.DataFrame.from_records(records)
    if df.empty:
        raise RuntimeError(f"{schema.layer}: no quantifiable rows after filtering")

    agg = _AGGREGATORS[schema.aggregator]
    return agg(df.groupby("gene", as_index=True, sort=False))


def load_omics_matrices() -> dict[str, pd.DataFrame]:
    """Load pr, ps, py matrices keyed by layer name. Column headers are
    canonical animal IDs; all paths resolved via ``alz.config``."""
    return {name: load_gene_matrix(schema) for name, schema in OMICS_SCHEMAS.items()}


def transcript_animal_canon_map(animal_meta: pd.DataFrame) -> dict[str, str]:
    """Map canonical animal ID → transcript-side animal_id.

    transcript animal_id form: ``D092_ma_6mo_ApTt``; canonical: ``D92``.
    """
    out: dict[str, str] = {}
    for tid in animal_meta["animal_id"]:
        canon = canonical_animal_id(tid.split("_", 1)[0])
        if canon is None:
            raise RuntimeError(f"could not canonicalize transcript id {tid!r}")
        if canon in out:
            raise RuntimeError(
                f"duplicate canonical animal {canon} in transcript meta "
                f"({out[canon]} and {tid})"
            )
        out[canon] = tid
    return out


def intersect_and_rekey(
    omics: dict[str, pd.DataFrame], transcript_map: dict[str, str]
) -> tuple[dict[str, pd.DataFrame], list[str], list[str]]:
    """Intersect omics animal coverage with transcript animal set.

    Returns (rekeyed_omics, kept_transcript_ids, dropped_transcript_ids):
        - rekeyed_omics: each matrix restricted to intersect animals and with
          columns renamed from canonical → transcript-side animal_id.
        - kept_transcript_ids: ordered list of transcript animal_ids present
          in all 4 layers (transcript + pr + ps + py).
        - dropped_transcript_ids: transcript animals missing from ≥1 omics layer.
    """
    canon_in_each: list[set[str]] = [set(df.columns) for df in omics.values()]
    all_omics_canon = set.intersection(*canon_in_each) if canon_in_each else set()

    transcript_canon = set(transcript_map.keys())
    intersect_canon = transcript_canon & all_omics_canon
    dropped_canon = transcript_canon - intersect_canon

    kept_transcript_ids = sorted(transcript_map[c] for c in intersect_canon)
    dropped_transcript_ids = sorted(transcript_map[c] for c in dropped_canon)

    rekeyed: dict[str, pd.DataFrame] = {}
    for layer, df in omics.items():
        sub = df.loc[:, sorted(intersect_canon)]
        sub.columns = [transcript_map[c] for c in sub.columns]
        sub = sub.reindex(columns=kept_transcript_ids)
        rekeyed[layer] = sub

    return rekeyed, kept_transcript_ids, dropped_transcript_ids
